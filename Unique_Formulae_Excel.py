import os
import re
import argparse
from openpyxl import load_workbook

# ----------------------------------------------------------------------------------------------------------------------
# --- Building Command ---

# Defining the argument parser for the command
parser = argparse.ArgumentParser()

# Excel workbook name
parser.add_argument('--filename', '-f',
                    help='(str) Name of excel workbook containing the formulae.',
                    type=str,
                    required=True)

# Separator of the formular names
parser.add_argument('--sep', '-s',
                    help="""(str) The separator of the formulae names. 
                    If none is give they will be printed in a new line each.""")

# Parse arguments to extract user inputs
args = parser.parse_args()
# ----------------------------------------------------------------------------------------------------------------------


# ----------------------------------------------------------------------------------------------------------------------
# --- Excel Workbook Path ---

# File directory
DIR = './'

# File name
FILENAME = args.filename

# File absolute path
FILEPATH = os.path.join(DIR, FILENAME)
# ----------------------------------------------------------------------------------------------------------------------


# ----------------------------------------------------------------------------------------------------------------------
# --- Extract Workbook formulae ---

# Loading workbook and ensuring that formulae are loaded
wb = load_workbook(FILEPATH, data_only=False, read_only=True)

# Excel formulae RegEx
pattern = r'[A-Z][A-Z0-9.]*(?=\()'

# Explanation:
# 1. [A-Z]: Start with a capital letter
# 2. [A-Z0-9_.]*: Zero or more of uppercase letters, digits, or dot . (cases like STDEV.S)
# (?=\(): A lookahead that asserts the next character is (, but does not include it in the match

# Set for storing unique formulae extracted from the whole workbook
unique_formulae = set()

# Search for formulae in each sheet
for sheet in wb.sheetnames:

    # Fetch worksheet
    ws = wb[sheet]

    # Search through all rows in the sheet
    for row in ws.iter_rows():

        # Search through all cells in the sheet
        for cell in row:

            # Check if cell has formula (all cells that start with '=...')
            if cell.data_type == 'f':

                # Trying to extract the formulae from text but TypeError is returned when instead of text
                # cell.value returns openpyxl.worksheet.formula.ArrayFormula object
                try:
                    # If this runs without problem cell.value has the formula
                    found_formula_names = re.findall(pattern, cell.value)
                except TypeError:
                    # If the above raises TypeError then cell.value.text has the formula
                    found_formula_names = re.findall(pattern, cell.value.text)

                # Check if formulae are found
                if found_formula_names:

                    # For each formula remove not needed prefixes
                    for f in found_formula_names:

                        current_formula = f

                        # Remove '_xlfn.' prefix, if present
                        if '_xlfn.' in current_formula:
                            current_formula.replace('_xlfn.', '')

                        # Add formula to set
                        unique_formulae.add(current_formula)

# Convert to list
unique_formulae = list(unique_formulae)

# Sort unique formulae
unique_formulae.sort()

# Add separator or print in new line if none is given
if args.sep:
    result = f'{args.sep}'.join(unique_formulae)
    print(result)
else:
    for f in unique_formulae:
        print(f)
